from openpyxl import load_workbook
'''
name="gdsa"
v_no=325
amt=33
rfid="737389nxzns"
'''
def xl_write(name,v_no,amt,rfid):

    book = load_workbook("Toll_Tax_Excel.xlsx")
    sheet = book['Sheet1']

    row=sheet.max_row
    col=sheet.max_column
    print(row,col)

    for i in range(row+1,row+2):
        for j in range(col):
            sheet['A'+str(i)]=name
            sheet['B'+str(i)]=amt
            sheet['C'+str(i)]=v_no
            sheet['D'+str(i)]=rfid
    book.save("Toll_Tax_Excel.xlsx") 

#xl_write(name,v_no,amt,rfid)

def xl_read():
    
    book = load_workbook("Toll_Tax_Excel.xlsx")
    sheet = book['Sheet1']

    row=sheet.max_row
    col=sheet.max_column
    a=[]

    for i in range(1,row+1):
        v=sheet.cell(row=i,column=4).value
        a.append(v)
    return a
        
